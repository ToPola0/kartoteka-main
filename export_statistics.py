"""
Moduł do eksportowania statystyk do plików Excel.
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from tkinter import filedialog, messagebox


def export_statistics_to_excel(statistics_obj, default_name="statystyki_kartoteka.xlsx"):
    """Eksportuje statystyki do pliku Excel z formatowaniem."""
    
    save_path = filedialog.asksaveasfilename(
        title="Zapisz statystyki do Excel",
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
        initialfile=default_name
    )
    
    if not save_path:
        return False
    
    try:
        summary = statistics_obj.get_summary()
        
        # Przygotuj dane do różnych arkuszy
        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            
            # Arkusz 1: Podsumowanie
            summary_data = [
                ["Kategoria", "Wartość"],
                ["Wszystkie osoby", summary['total_people']],
                ["Kobiety", summary['total_females']],
                ["Mężczyźni", summary['total_males']],
                ["Przeskanowane pliki", summary['files_scanned']],
                ["Przeskanowane arkusze", summary['sheets_scanned']],
                ["Unikalne adresy", summary['unique_addresses']],
                ["Błędy", summary['errors_count']],
                ["Ostrzeżenia", summary['warnings_count']],
                ["Nieznane imiona", summary['unknown_names_count']],
                ["Jubileusze", summary['jubilees_count']],
                ["Śluby w zakresie", summary['marriages_in_range_count']],
                ["Czas analizy (s)", f"{summary['analysis_duration']:.2f}"],
            ]
            df_summary = pd.DataFrame(summary_data[1:], columns=summary_data[0])
            df_summary.to_excel(writer, sheet_name='Podsumowanie', index=False)
            
            # Arkusz 2: Statystyki wieku
            age_data = [
                ["Statystyka", "Wartość"],
                ["Średnia wieku", f"{summary['age_average']:.1f} lat"],
                ["Mediana wieku", f"{summary['age_median']:.1f} lat"],
                ["Najmłodszy", f"{summary['age_min']} lat"],
                ["Najstarszy", f"{summary['age_max']} lat"],
                ["Rozstęp wieku", f"{summary['age_max'] - summary['age_min']} lat"],
            ]
            df_age = pd.DataFrame(age_data[1:], columns=age_data[0])
            df_age.to_excel(writer, sheet_name='Statystyki wieku', index=False)
            
            # Arkusz 3: Grupy wiekowe
            age_groups_data = [["Grupa wiekowa", "Liczba osób", "Procent"]]
            for group, count in summary['age_groups'].items():
                percentage = (count / summary['total_people'] * 100) if summary['total_people'] > 0 else 0
                age_groups_data.append([group, count, f"{percentage:.1f}%"])
            df_age_groups = pd.DataFrame(age_groups_data[1:], columns=age_groups_data[0])
            df_age_groups.to_excel(writer, sheet_name='Grupy wiekowe', index=False)
            
            # Arkusz 4: Urodziny w dekadach
            if summary['birth_decades']:
                birth_decades_data = [["Dekada", "Liczba urodzin", "Procent"]]
                sorted_decades = sorted(summary['birth_decades'].items())
                for decade, count in sorted_decades:
                    percentage = (count / summary['total_people'] * 100) if summary['total_people'] > 0 else 0
                    birth_decades_data.append([f"{decade}s", count, f"{percentage:.1f}%"])
                df_birth = pd.DataFrame(birth_decades_data[1:], columns=birth_decades_data[0])
                df_birth.to_excel(writer, sheet_name='Urodziny w dekadach', index=False)
            
            # Arkusz 5: Śluby w dekadach
            if summary['marriage_decades']:
                marriage_decades_data = [["Dekada", "Liczba ślubów", "Procent"]]
                sorted_decades = sorted(summary['marriage_decades'].items())
                total_marriages = sum(summary['marriage_decades'].values())
                for decade, count in sorted_decades:
                    percentage = (count / total_marriages * 100) if total_marriages > 0 else 0
                    marriage_decades_data.append([f"{decade}s", count, f"{percentage:.1f}%"])
                df_marriage = pd.DataFrame(marriage_decades_data[1:], columns=marriage_decades_data[0])
                df_marriage.to_excel(writer, sheet_name='Śluby w dekadach', index=False)
            
            # Formatowanie i auto-dopasowanie kolumn
            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Auto-dopasowanie szerokości kolumn
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Formatowanie nagłówków
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=12)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                
                # Formatowanie zawartości
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        cell.border = border
        
        messagebox.showinfo("Sukces", f"Statystyki zapisane do:\n{save_path}")
        return True
        
    except Exception as e:
        messagebox.showerror("Błąd", f"Nie udało się zapisać statystyk:\n{e}")
        return False


def export_all_results_to_excel(found_people, statistics_obj, jubilees_found=None, marriages_in_range=None, all_unknown=None, default_name="wszystkie_wyniki.xlsx"):
    """Eksportuje wszystkie wyniki (osoby + statystyki + jubileusze + śluby + nieznane) do jednego pliku Excel."""
    
    save_path = filedialog.asksaveasfilename(
        title="Zapisz wszystkie wyniki do Excel",
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
        initialfile=default_name
    )
    
    if not save_path:
        return False
    
    try:
        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            
            # Arkusz 1: Znalezione osoby - WSZYSTKIE POLA
            if found_people:
                people_data = []
                from data_processing import format_person_name
                for person in found_people:
                    imie = format_person_name(person.get("imie", ""))
                    nazwisko = format_person_name(person.get("nazwisko", ""))
                    adres = person.get("adres", "")
                    old_address = person.get("old_address", "")
                    wiek = person.get("wiek", "")
                    plec = "Kobieta" if person.get("plec", "") == "K" else "Mężczyzna" if person.get("plec", "") == "M" else ""
                    plik = person.get("file", "")
                    file_path = person.get("file_path", "")
                    people_data.append([imie, nazwisko, adres, old_address, wiek, plec, plik, file_path])
                
                df_people = pd.DataFrame(people_data, columns=[
                    "Imię", "Nazwisko", "Adres aktualny", "Adres stary", "Wiek", "Płeć", "Plik źródłowy", "Pełna ścieżka"
                ])
                df_people.to_excel(writer, sheet_name='Znalezione osoby', index=False)
            
            # Arkusz 2-6: Statystyki (jak w poprzedniej funkcji)
            summary = statistics_obj.get_summary()
            
            # Arkusz 2: Podsumowanie (podstawowe liczby)
            summary_data = [
                ["Kategoria", "Wartość"],
                ["Wszystkie osoby", summary['total_people']],
                ["Kobiety", summary['total_females']],
                ["Mężczyźni", summary['total_males']],
                ["Przeskanowane pliki", summary['files_scanned']],
                ["Przeskanowane arkusze", summary['sheets_scanned']],
                ["Unikalne adresy", summary['unique_addresses']],
                ["Błędy", summary['errors_count']],
                ["Ostrzeżenia", summary['warnings_count']],
                ["Nieznane imiona", summary['unknown_names_count']],
                ["Jubileusze", summary['jubilees_count']],
                ["Śluby w zakresie", summary['marriages_in_range_count']],
                ["Czas analizy (s)", f"{summary['analysis_duration']:.2f}"],
            ]
            df_summary = pd.DataFrame(summary_data[1:], columns=summary_data[0])
            df_summary.to_excel(writer, sheet_name='Podsumowanie', index=False)
            
            # Arkusz 3: Statystyki wieku (szczegółowe)
            age_data = [
                ["Statystyka", "Wartość"],
                ["Średnia wieku", f"{summary['age_average']:.1f} lat"],
                ["Mediana wieku", f"{summary['age_median']:.1f} lat"],
                ["Najmłodszy", f"{summary['age_min']} lat"],
                ["Najstarszy", f"{summary['age_max']} lat"],
                ["Rozstęp wieku", f"{summary['age_max'] - summary['age_min']} lat"],
            ]
            df_age = pd.DataFrame(age_data[1:], columns=age_data[0])
            df_age.to_excel(writer, sheet_name='Statystyki wieku', index=False)
            
            # Arkusz 4: Grupy wiekowe z procentami
            age_groups_data = [["Grupa wiekowa", "Liczba osób", "Procent"]]
            for group, count in summary['age_groups'].items():
                percentage = (count / summary['total_people'] * 100) if summary['total_people'] > 0 else 0
                age_groups_data.append([group, count, f"{percentage:.1f}%"])
            df_age_groups = pd.DataFrame(age_groups_data[1:], columns=age_groups_data[0])
            df_age_groups.to_excel(writer, sheet_name='Grupy wiekowe', index=False)
            
            # Arkusz 5: Urodziny w dekadach z procentami
            if summary['birth_decades']:
                birth_decades_data = [["Dekada", "Liczba urodzin", "Procent"]]
                sorted_decades = sorted(summary['birth_decades'].items())
                for decade, count in sorted_decades:
                    percentage = (count / summary['total_people'] * 100) if summary['total_people'] > 0 else 0
                    birth_decades_data.append([f"{decade}s", count, f"{percentage:.1f}%"])
                df_birth = pd.DataFrame(birth_decades_data[1:], columns=birth_decades_data[0])
                df_birth.to_excel(writer, sheet_name='Urodziny w dekadach', index=False)
            
            # Arkusz 6: Śluby w dekadach z procentami
            if summary['marriage_decades']:
                marriage_decades_data = [["Dekada", "Liczba ślubów", "Procent"]]
                sorted_decades = sorted(summary['marriage_decades'].items())
                total_marriages = sum(summary['marriage_decades'].values())
                for decade, count in sorted_decades:
                    percentage = (count / total_marriages * 100) if total_marriages > 0 else 0
                    marriage_decades_data.append([f"{decade}s", count, f"{percentage:.1f}%"])
                df_marriage = pd.DataFrame(marriage_decades_data[1:], columns=marriage_decades_data[0])
                df_marriage.to_excel(writer, sheet_name='Śluby w dekadach', index=False)
            
            # Arkusz: Jubileusze (jeśli dostępne)
            if jubilees_found and len(jubilees_found) > 0:
                jubilees_data = []
                for j in sorted(jubilees_found, key=lambda x: x.get("days", 0)):
                    jubilees_data.append([
                        j.get("date", ""),
                        j.get("years", ""),
                        j.get("husband", ""),
                        j.get("wife", ""),
                        j.get("surname", ""),
                        j.get("old_address", ""),
                        j.get("type", "MAŁŻONKOWIE"),
                        j.get("days", "")
                    ])
                df_jubilees = pd.DataFrame(jubilees_data, columns=[
                    "Data jubileuszu", "Lata małżeństwa", "Mąż", "Żona", "Nazwisko", "Stary adres", "Typ", "Dni do jubileuszu"
                ])
                df_jubilees.to_excel(writer, sheet_name='Jubileusze', index=False)
            
            # Arkusz: Śluby w zakresie lat (jeśli dostępne)
            if marriages_in_range and len(marriages_in_range) > 0:
                marriages_data = []
                for m in sorted(marriages_in_range, key=lambda x: x.get("year", 0)):
                    marriages_data.append([
                        m.get("year", ""),
                        m.get("date", ""),
                        m.get("husband", ""),
                        m.get("wife", ""),
                        m.get("surname", ""),
                        m.get("address", ""),
                        m.get("old_address", ""),
                        m.get("type", ""),
                        m.get("file_path", "")
                    ])
                df_marriages = pd.DataFrame(marriages_data, columns=[
                    "Rok", "Data ślubu", "Mąż", "Żona", "Nazwisko", "Adres", "Stary adres", "Typ", "Plik źródłowy"
                ])
                df_marriages.to_excel(writer, sheet_name='Śluby w zakresie lat', index=False)
            
            # Arkusz: Nieznane imiona (jeśli dostępne)
            if all_unknown and len(all_unknown) > 0:
                unknown_data = []
                for name, locations in sorted(all_unknown.items()):
                    for location in locations:
                        unknown_data.append([name, location, len(locations)])
                df_unknown = pd.DataFrame(unknown_data, columns=[
                    "Nieznane imię", "Lokalizacja", "Liczba wystąpień"
                ])
                df_unknown.to_excel(writer, sheet_name='Nieznane imiona', index=False)
            
            # Formatowanie wszystkich arkuszy
            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Auto-dopasowanie szerokości kolumn
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Formatowanie nagłówków
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=12)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                
                # Formatowanie zawartości
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        cell.border = border
        
        messagebox.showinfo("Sukces", f"Wszystkie wyniki zapisane do:\n{save_path}")
        return True
        
    except Exception as e:
        messagebox.showerror("Błąd", f"Nie udało się zapisać wyników:\n{e}")
        return False
