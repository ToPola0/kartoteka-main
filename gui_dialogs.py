import tkinter as tk
from tkinter import scrolledtext, messagebox, filedialog
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import logging
import os
import re
from datetime import datetime
from data_processing import format_person_name, remove_diacritics, extract_number_from_text
from file_operations import save_names_to_json
from analysis import extract_marriage_info, extract_grandparents_marriage_info

def open_excel_file_for_editing(selected_name, locations, folder_path):
    """Otwiera okno do edycji pliku Excel dla wybranego imienia."""
    if not locations:
        messagebox.showinfo("Brak lokalizacji", "Nie znaleziono lokalizacji dla wybranego imienia.")
        return

    edit_window = tk.Toplevel()
    edit_window.title(f"Edytuj plik dla imienia: {selected_name}")
    edit_window.geometry("600x400")

    tk.Label(edit_window, text=f"Imię: {selected_name}").pack(pady=5)
    tk.Label(edit_window, text="Wybierz plik i arkusz do edycji:").pack(pady=5)

    location_listbox = tk.Listbox(edit_window, width=80, height=15)
    location_listbox.pack(padx=10, pady=5)

    for location in locations:
        location_listbox.insert(tk.END, location)

    def open_selected_file():
        selected_index = location_listbox.curselection()
        if not selected_index:
            messagebox.showwarning("Błąd", "Nie wybrano pliku do edycji.")
            return

        selected_location = location_listbox.get(selected_index)
        file_path, sheet_name = selected_location.split(" -> ")

        try:
            xl = pd.ExcelFile(file_path)
            df = xl.parse(sheet_name, header=None)

            edit_excel_window = tk.Toplevel(edit_window)
            edit_excel_window.title(f"Edycja: {file_path} -> {sheet_name}")
            edit_excel_window.geometry("800x600")

            tk.Label(edit_excel_window, text=f"Edytujesz: {file_path} -> {sheet_name}").pack(pady=5)

            text_area = scrolledtext.ScrolledText(edit_excel_window, width=100, height=30)
            text_area.pack(padx=10, pady=10)

            for row in df.itertuples(index=False):
                text_area.insert(tk.END, "\t".join(map(str, row)) + "\n")

            def save_changes_to_excel():
                modified_data = text_area.get("1.0", tk.END).strip().split("\n")
                new_data = [line.split("\t") for line in modified_data]
                new_df = pd.DataFrame(new_data)
                with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                    new_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                messagebox.showinfo("Sukces", f"Zmiany zapisano w pliku: {file_path}, arkusz: {sheet_name}")
                edit_excel_window.destroy()

            tk.Button(edit_excel_window, text="Zapisz zmiany", command=save_changes_to_excel).pack(pady=10)

        except Exception as e:
            messagebox.showerror("Błąd", f"Nie udało się otworzyć pliku: {e}")

    tk.Button(edit_window, text="Otwórz plik do edycji", command=open_selected_file).pack(pady=10)
    tk.Button(edit_window, text="Zamknij", command=edit_window.destroy).pack(pady=5)

def edit_unknown_name(all_unknown, names_dict, folder_path, result_text):
    """Okno do edycji nieznanych imion."""
    if not all_unknown:
        messagebox.showinfo("Brak nieznanych imion", "Nie znaleziono nieznanych imion do edycji.")
        return

    edit_window = tk.Toplevel()
    edit_window.title("Edytuj nieznane imię")
    edit_window.geometry("800x550")
    edit_window.configure(bg="#ECF0F1")
    
    # Nagłówek
    header = tk.Frame(edit_window, bg="#2C3E50", height=60)
    header.pack(fill="x")
    header.pack_propagate(False)
    
    title_label = tk.Label(header, text="✏️ Edycja nieznanych imion",
                          font=("Segoe UI", 16, "bold"),
                          bg="#2C3E50", fg="white")
    title_label.pack(pady=15)
    
    # Główna zawartość
    content = tk.Frame(edit_window, bg="#ECF0F1")
    content.pack(fill="both", expand=True, padx=15, pady=15)

    # Lewy panel - lista nieznanych imion
    left_frame = tk.Frame(content, bg="#ECF0F1")
    left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
    
    tk.Label(left_frame, text="Wybierz imię do poprawy:",
            font=("Segoe UI", 10, "bold"),
            bg="#ECF0F1", fg="#2C3E50").pack(pady=(0, 8))
    
    # Lista z scrollbarem
    list_frame = tk.Frame(left_frame, bg="white", relief="solid", bd=1)
    list_frame.pack(fill=tk.BOTH, expand=True)
    
    scrollbar = tk.Scrollbar(list_frame, orient="vertical")
    unknown_listbox = tk.Listbox(list_frame, 
                                 width=45, 
                                 height=20,
                                 font=("Segoe UI", 10),
                                 bg="white",
                                 fg="#2C3E50",
                                 relief="flat",
                                 selectmode=tk.SINGLE,
                                 yscrollcommand=scrollbar.set)
    scrollbar.config(command=unknown_listbox.yview)
    
    unknown_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    for name in list(all_unknown.keys()):
        unknown_listbox.insert(tk.END, name)

    # Prawy panel - formularz edycji
    right_frame = tk.Frame(content, bg="#ECF0F1", width=280)
    right_frame.pack(side=tk.RIGHT, fill=tk.Y)
    right_frame.pack_propagate(False)
    
    tk.Label(right_frame, text="Wprowadź poprawioną wersję:",
            font=("Segoe UI", 10, "bold"),
            bg="#ECF0F1", fg="#2C3E50").pack(pady=(0, 5))
    
    corrected_name_entry = tk.Entry(right_frame, 
                                     width=30,
                                     font=("Segoe UI", 10),
                                     relief="solid",
                                     bd=1)
    corrected_name_entry.pack(pady=(0, 15), fill="x")

    tk.Label(right_frame, text="Płeć (przy dodaniu do bazy):",
            font=("Segoe UI", 10, "bold"),
            bg="#ECF0F1", fg="#2C3E50").pack(pady=(0, 5))
    
    gender_var = tk.StringVar(value="K")
    rb_frame = tk.Frame(right_frame, bg="#ECF0F1")
    rb_frame.pack(pady=(0, 20))
    
    tk.Radiobutton(rb_frame, text="Kobieta (K)", 
                  variable=gender_var, 
                  value="K",
                  font=("Segoe UI", 9),
                  bg="#ECF0F1",
                  activebackground="#ECF0F1").pack(anchor="w", pady=2)
    tk.Radiobutton(rb_frame, text="Mężczyzna (M)", 
                  variable=gender_var, 
                  value="M",
                  font=("Segoe UI", 9),
                  bg="#ECF0F1",
                  activebackground="#ECF0F1").pack(anchor="w", pady=2)

    # Separator
    sep = tk.Frame(right_frame, height=2, bg="#BDC3C7")
    sep.pack(fill="x", pady=(0, 15))
    
    info_label = tk.Label(right_frame, text="", 
                         fg="#27AE60",
                         font=("Segoe UI", 9),
                         bg="#ECF0F1",
                         wraplength=260)
    info_label.pack(pady=(0, 15))

    def apply_correction():
        selected_index = unknown_listbox.curselection()
        if not selected_index:
            messagebox.showwarning("Błąd", "Nie wybrano imienia do poprawy.")
            return
        selected_name = unknown_listbox.get(selected_index)
        corrected_name = corrected_name_entry.get().strip()
        if not corrected_name:
            messagebox.showwarning("Błąd", "Nie wprowadzono poprawionej wersji imienia.")
            return
        corrected_name_fmt = format_person_name(corrected_name)

        for location in list(all_unknown.get(selected_name, [])):
            file_part, sheet_name = location.split(" -> ")
            full_path = file_part if os.path.isabs(file_part) else os.path.join(folder_path, file_part)
            try:
                workbook = load_workbook(full_path)
                if sheet_name not in workbook.sheetnames:
                    result_text.insert(tk.END, f"[ERROR] Arkusz '{sheet_name}' nie istnieje w pliku: {full_path}\n")
                    continue
                sheet = workbook[sheet_name]
                replaced_any = False
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            parts = re.split(r'(\W+)', cell.value)
                            changed = False
                            for i, part in enumerate(parts):
                                if part.strip() == "":
                                    continue
                                if re.search(r'\w', part):
                                    if remove_diacritics(part).strip().lower() == selected_name:
                                        parts[i] = corrected_name_fmt
                                        changed = True
                            if changed:
                                cell.value = "".join(parts)
                                replaced_any = True
                if replaced_any:
                    workbook.save(full_path)
                    result_text.insert(tk.END, f"[INFO] Zaktualizowano imię '{selected_name}' na '{corrected_name_fmt}' w pliku: {full_path}, arkusz: {sheet_name}\n")
                else:
                    result_text.insert(tk.END, f"[INFO] Nie znaleziono dokładnych wystąpień '{selected_name}' w pliku: {full_path}, arkusz: {sheet_name}\n")
            except FileNotFoundError:
                result_text.insert(tk.END, f"[ERROR] Plik nie został znaleziony: {full_path}\n")
            except Exception as e:
                result_text.insert(tk.END, f"[ERROR] Nie udało się zaktualizować pliku {full_path}: {e}\n")

        if selected_name in all_unknown:
            del all_unknown[selected_name]
        unknown_listbox.delete(selected_index)
        corrected_name_entry.delete(0, tk.END)
        info_label.config(text=f"Imię '{selected_name}' poprawione.")
        messagebox.showinfo("Sukces", f"Imię '{selected_name}' zostało poprawione na '{corrected_name_fmt}'.")

    def add_to_database():
        selected_index = unknown_listbox.curselection()
        if not selected_index:
            messagebox.showwarning("Błąd", "Nie wybrano imienia do dodania.")
            return
        selected_name = unknown_listbox.get(selected_index)
        corrected_name = corrected_name_entry.get().strip()
        key_name = format_person_name(corrected_name) if corrected_name else format_person_name(selected_name)
        gender = gender_var.get()
        names_dict[remove_diacritics(key_name).strip().lower()] = gender

        # Domyślna ścieżka to folder z plikami Excel
        default_json_path = os.path.join(folder_path, "imiona.json") if folder_path else "imiona.json"
        
        save_path = filedialog.asksaveasfilename(
            title="Zapisz/wybierz plik JSON z imionami", 
            defaultextension=".json", 
            filetypes=[("JSON", "*.json")],
            initialdir=folder_path if folder_path else None,
            initialfile="imiona.json"
        )
        if save_path:
            if save_names_to_json({key_name: gender}, save_path):
                result_text.insert(tk.END, f"[INFO] Dodano imię '{key_name}' jako '{gender}' do pliku: {save_path}\n")
            else:
                result_text.insert(tk.END, f"[ERROR] Nie udało się zapisać do pliku JSON\n")
                return
        else:
            result_text.insert(tk.END, f"[INFO] Imię '{key_name}' dodano do bieżącej sesji (nie zapisano do pliku).\n")

        if selected_name in all_unknown:
            del all_unknown[selected_name]
        unknown_listbox.delete(selected_index)
        corrected_name_entry.delete(0, tk.END)
        info_label.config(text=f"Imię '{key_name}' dodano do bazy jako '{gender}'.")
        messagebox.showinfo("Sukces", f"Imię '{key_name}' dodano jako '{gender}' do bazy (pamięć sesji zaktualizowana).")

    # Przyciski akcji
    btn_style = {
        "font": ("Segoe UI", 10, "bold"),
        "relief": "flat",
        "bd": 0,
        "cursor": "hand2",
        "padx": 15,
        "pady": 10
    }
    
    btn_frame = tk.Frame(right_frame, bg="#ECF0F1")
    btn_frame.pack(fill=tk.X, pady=(5, 0))
    
    apply_btn = tk.Button(btn_frame, text="✅ Zastosuj poprawkę", 
                         command=apply_correction,
                         bg="#27AE60", 
                         fg="white",
                         activebackground="#229954",
                         activeforeground="white",
                         **btn_style)
    apply_btn.pack(fill=tk.X, pady=(0, 8))
    
    add_btn = tk.Button(btn_frame, text="➕ Dodaj do bazy imion", 
                       command=add_to_database,
                       bg="#3498DB",
                       fg="white",
                       activebackground="#2980B9",
                       activeforeground="white",
                       **btn_style)
    add_btn.pack(fill=tk.X, pady=(0, 8))
    
    close_btn = tk.Button(btn_frame, text="❌ Zamknij", 
                         command=edit_window.destroy,
                         bg="#95A5A6",
                         fg="white",
                         activebackground="#7F8C8D",
                         activeforeground="white",
                         **btn_style)
    close_btn.pack(fill=tk.X)

def show_results_dialog(found_people, root):
    """Wyświetla okno z wynikami analizy."""
    dialog = tk.Toplevel(root)
    from config import set_window_icon
    set_window_icon(dialog)
    dialog.title("Analiza zakończona - wyniki")
    dialog.geometry("750x500")
    dialog.configure(bg="#f0f0f0")


    tk.Label(dialog, text="Wyniki, które zostaną zapisane do pliku XLSX:", font=("Arial", 12, "bold"), bg="#f0f0f0", fg="#333").pack(pady=10)

    # PANEL FILTROWANIA (na dole, jedno pole)
    filter_frame = tk.Frame(dialog, bg="#e0e0e0")
    # Przeniesiemy pack() na dół, tuż nad btn_frame
    filter_var = tk.StringVar()
    tk.Label(filter_frame, text="Filtr imię/nazwisko:", bg="#e0e0e0").pack(side="left", padx=(4,2))
    filter_entry = tk.Entry(filter_frame, textvariable=filter_var, width=28)
    filter_entry.pack(side="left", padx=(0,8))
    
    # WYNIKI
    results_area = scrolledtext.ScrolledText(dialog, width=95, height=22, bg="#ffffff", fg="#333", font=("Courier", 10), relief="flat", bd=1, cursor="arrow")
    results_area.pack(fill="both", expand=True, padx=15, pady=5)
    results_area.tag_configure("result_name", font=("Courier", 10, "bold"))
    results_area.tag_configure("result_link", font=("Courier", 10, "bold"), foreground="blue", underline=True)

    sort_map = {
        "Brak": None,
        "Wiek rosnąco": "wiek",
        "Adres rosnąco": "adres",
        "Stary adres rosnąco": "stary_adres",
        "Stary numer rosnąco": "stary_numer",
        "Alfabetycznie (nazwisko, imię)": "alfabetycznie"
    }
    sort_var = tk.StringVar(value="Stary numer rosnąco")

    # Panel filtrowania na dole, tuż nad przyciskami
    filter_frame.pack(fill="x", padx=15, pady=(0, 5))

    btn_frame = tk.Frame(dialog, bg="#f0f0f0")
    btn_frame.pack(pady=10)
    tk.Label(btn_frame, text="Sortuj:", font=("Arial", 11, "bold"), bg="#f0f0f0", fg="#333").pack(side=tk.LEFT, padx=10)
    tk.OptionMenu(btn_frame, sort_var, *sort_map.keys()).pack(side=tk.LEFT, padx=10)


    def update_results_area():
        results_area.configure(state="normal")
        results_area.delete("1.0", tk.END)
        # FILTROWANIE (tylko początek imienia lub nazwiska)
        filter_val = filter_var.get().strip().lower()
        def match_start(val, text):
            return text.startswith(val) if val else True
        list_to_display = [p for p in found_people if not filter_val or match_start(filter_val, format_person_name(p.get("imie", "")).lower()) or match_start(filter_val, format_person_name(p.get("nazwisko", "")).lower())]
        sort_key = sort_map.get(sort_var.get())
        try:
            if sort_key:
                if sort_key == "wiek":
                    list_to_display.sort(key=lambda p: (p.get("wiek") is None, p.get("wiek")))
                elif sort_key == "adres":
                    list_to_display.sort(key=lambda p: ((p.get("adres") or "").lower()))
                elif sort_key == "stary_adres":
                    list_to_display.sort(key=lambda p: ((p.get("old_address") or "").lower()))
                elif sort_key == "stary_numer":
                    list_to_display.sort(key=lambda p: (
                        (p.get("old_address") is None),
                        (extract_number_from_text(p.get("old_address")) if extract_number_from_text(p.get("old_address")) is not None else float('inf'))
                    ))
                elif sort_key == "alfabetycznie":
                    list_to_display.sort(key=lambda p: (
                        (format_person_name(p.get("nazwisko","")).lower()),
                        (format_person_name(p.get("imie","")).lower())
                    ))
        except (KeyError, AttributeError, TypeError) as e:
            logging.warning(f"Błąd sortowania w dialogu wyników: {e}")

        if list_to_display:
            person_counter = 0
            if any(p.get("file_path") for p in list_to_display):
                results_area.insert(tk.END, "Kliknij imię i nazwisko aby otworzyć kartotekę\n", "result_name")
                results_area.insert(tk.END, "-" * 50 + "\n\n")
            for p in list_to_display:
                person_counter += 1
                imie = format_person_name(p.get("imie", ""))
                nazwisko = format_person_name(p.get("nazwisko", ""))
                adres = p.get("adres", "") or ""
                if p.get("old_address"):
                    adres += f" (stary: {p['old_address']})"
                wiek = p.get("wiek", "")
                file_path = p.get("file_path", "")
                tag_name = f"person_link_{person_counter}"
                results_area.insert(tk.END, f"{nazwisko} {imie}", tag_name)
                results_area.tag_config(tag_name, foreground="blue", underline=True)
                if file_path and os.path.exists(file_path):
                    results_area.tag_bind(tag_name, "<Button-1>", lambda event, path=file_path: os.startfile(path))
                    results_area.tag_bind(tag_name, "<Enter>", lambda event, t=tag_name, area=results_area: area.tag_config(t, font=("Arial", 10, "underline")) if area.winfo_exists() else None)
                    results_area.tag_bind(tag_name, "<Leave>", lambda event, t=tag_name, area=results_area: area.tag_config(t, font=("Arial", 10)) if area.winfo_exists() else None)
                results_area.insert(tk.END, f" - {adres} - wiek: {wiek}\n")

            # Dodaj informację o przypisanych medianach na końcu listy
            median_count = sum(1 for p in list_to_display if p.get("wiek_info") == "MEDIANA_WIEKU" or (isinstance(p.get("wiek"), str) and p.get("wiek") == "MEDIANA_WIEKU"))
            if median_count > 0:
                results_area.insert(tk.END, f"\n[INFO] Liczba osób z przypisanym wiekiem na podstawie mediany populacji: {median_count}\n")
        else:
            results_area.insert(tk.END, "Brak wyników do zapisania.\n")
        results_area.configure(state="disabled")

    sort_var.trace_add("write", lambda *args: update_results_area())
    filter_entry.bind("<Return>", lambda e: update_results_area())
    filter_btn = tk.Button(filter_frame, text="Filtruj", command=update_results_area)
    filter_btn.pack(side="left", padx=(2,4))
    update_results_area()

    from gui_main import save_found_people_to_xlsx
    tk.Button(btn_frame, text="Zapisz do pliku", command=lambda: save_found_people_to_xlsx(found_people, sort_map.get(sort_var.get())), bg="#607D8B", fg="white", font=("Arial", 11, "bold"), relief="raised", bd=2).pack(side=tk.LEFT, padx=10)
    tk.Button(btn_frame, text="Zamknij", command=dialog.destroy, bg="#9E9E9E", fg="white", font=("Arial", 11, "bold"), relief="raised", bd=2).pack(side=tk.LEFT, padx=10)

def show_marriages_dialog(folder_path):
    """Wyświetla okno z wyszukiwaniem ślubów w zakresie lat."""
    if not folder_path or not os.path.exists(folder_path):
        messagebox.showwarning("Błąd", "Najpierw wybierz folder z kartotek!")
        return
    
    dialog = tk.Toplevel()
    from config import set_window_icon
    set_window_icon(dialog)
    dialog.title("Wyszukiwanie ślubów w latach")
    dialog.geometry("900x650")
    dialog.configure(bg="#ECF0F1")
    
    # Nagłówek
    header = tk.Frame(dialog, bg="#2C3E50", height=60)
    header.pack(fill="x")
    header.pack_propagate(False)
    
    title_label = tk.Label(header, text="💍 Wyszukiwanie ślubów w zakresie lat",
                          font=("Segoe UI", 16, "bold"),
                          bg="#2C3E50", fg="white")
    title_label.pack(pady=15)
    
    # Zawartość
    content = tk.Frame(dialog, bg="#ECF0F1")
    content.pack(fill="both", expand=True, padx=15, pady=15)
    
    # Panel wyszukiwania
    search_frame = tk.Frame(content, bg="#ECF0F1")
    search_frame.pack(fill="x", pady=(0, 10))
    
    tk.Label(search_frame, text="Zakres lat ślubów:", 
            font=("Segoe UI", 11, "bold"),
            bg="#ECF0F1", fg="#2C3E50").grid(row=0, column=0, sticky="w", padx=(0, 10))
    
    tk.Label(search_frame, text="od:", 
            font=("Segoe UI", 10),
            bg="#ECF0F1", fg="#2C3E50").grid(row=0, column=1, padx=(0, 5))
    
    year_from_var = tk.IntVar(value=1900)
    year_from_entry = tk.Entry(search_frame, textvariable=year_from_var, width=8,
                               font=("Segoe UI", 10))
    year_from_entry.grid(row=0, column=2, padx=(0, 15))
    
    tk.Label(search_frame, text="do:", 
            font=("Segoe UI", 10),
            bg="#ECF0F1", fg="#2C3E50").grid(row=0, column=3, padx=(0, 5))
    
    year_to_var = tk.IntVar(value=datetime.now().year)
    year_to_entry = tk.Entry(search_frame, textvariable=year_to_var, width=8,
                             font=("Segoe UI", 10))
    year_to_entry.grid(row=0, column=4, padx=(0, 15))
    
    search_btn = tk.Button(search_frame, text="🔍 Szukaj", 
                          font=("Segoe UI", 10, "bold"),
                          bg="#3498DB", fg="white",
                          relief="flat", padx=15, pady=5,
                          cursor="hand2")
    search_btn.grid(row=0, column=5)
    
    # Obszar wyników
    results_frame = tk.Frame(content, bg="#FFFFFF", relief="solid", bd=1)
    results_frame.pack(fill="both", expand=True, pady=(10, 0))
    
    results_area = scrolledtext.ScrolledText(results_frame, width=100, height=25, 
                                            bg="#FFFFFF", fg="#333", 
                                            font=("Courier", 10), 
                                            relief="flat", wrap=tk.WORD)
    results_area.pack(fill="both", expand=True, padx=5, pady=5)
    
    # Zmienna do przechowywania znalezionych ślubów
    marriages_found = []
    
    def save_marriages_to_excel():
        """Zapisuje znalezione śluby do pliku Excel."""
        if not marriages_found:
            messagebox.showinfo("Brak danych", "Brak ślubów do zapisania. Najpierw wykonaj wyszukiwanie.")
            return
        
        from tkinter import filedialog
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Plik Excel", "*.xlsx")],
            initialfile=f"Sluby_{year_from_var.get()}-{year_to_var.get()}.xlsx"
        )
        
        if not save_path:
            return
        
        try:
            # Przygotuj dane do zapisu
            data = []
            for m in sorted(marriages_found, key=lambda x: x["year"]):
                type_str = m.get("type", "MAŁŻONKOWIE")
                old = m.get("old_address", "")
                addr_info = m.get("address", "")
                location = f"{addr_info}" + (f" (stary: {old})" if old else "")
                
                data.append({
                    "Data ślubu": m["date"],
                    "Rok": m["year"],
                    "Nazwisko": m["surname"],
                    "Mąż": m["husband"],
                    "Żona": m["wife"],
                    "Adres": location,
                    "Typ": type_str,
                    "Plik": os.path.basename(m.get("file_path", ""))
                })
            
            # Zapisz do Excel
            df = pd.DataFrame(data)
            df.to_excel(save_path, index=False, engine='openpyxl')
            
            # Dopasuj szerokości kolumn
            from openpyxl import load_workbook
            wb = load_workbook(save_path)
            ws = wb.active
            
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Max 50 znaków
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(save_path)
            
            messagebox.showinfo("Sukces", f"Zapisano {len(marriages_found)} ślubów do pliku:\n{save_path}")
        except Exception as e:
            messagebox.showerror("Błąd", f"Nie udało się zapisać pliku:\n{e}")
    
    def search_marriages():
        """Szuka ślubów w podanym zakresie lat."""
        nonlocal marriages_found
        try:
            year_from = year_from_var.get()
            year_to = year_to_var.get()
            
            if year_from < 1800 or year_to > 2100 or year_from > year_to:
                messagebox.showwarning("Błąd", "Podano nieprawidłowy zakres lat!")
                return
            
            results_area.delete("1.0", tk.END)
            results_area.insert(tk.END, f"Szukam ślubów w latach {year_from}-{year_to}...\n\n")
            results_area.update_idletasks()
            
            marriages_found = []
            files_scanned = 0
            
            # Przeszukaj wszystkie pliki Excel w folderze
            for filename in os.listdir(folder_path):
                if not filename.endswith(('.xlsx', '.xls')) or filename.startswith('~$'):
                    continue
                
                file_path = os.path.join(folder_path, filename)
                files_scanned += 1
                
                try:
                    df = pd.read_excel(file_path, sheet_name=0, header=None)
                    
                    # Pobierz nazwisko z pliku
                    surname = filename.replace('.xlsx', '').replace('.xls', '')
                    
                    # Pobierz adres
                    address = ""
                    old_address = ""
                    try:
                        if df.shape[0] >= 4 and df.shape[1] >= 7:
                            address_parts = []
                            for col in range(5, 7):
                                column_data = df.iloc[0:4, col].dropna()
                                if not column_data.empty:
                                    address_parts.append(str(column_data.iloc[0]))
                            address = " ".join(address_parts)
                            
                            old_address_parts = []
                            for col in range(5, 7):
                                column_data = df.iloc[1:4, col].dropna()
                                if not column_data.empty:
                                    old_address_parts.append(str(column_data.iloc[0]))
                            old_address = " ".join(old_address_parts)
                    except Exception:
                        pass
                    
                    # Sprawdź śluby małżonków
                    marriage_info = extract_marriage_info(df)
                    if marriage_info.get("marriage_date"):
                        try:
                            marriage_year = datetime.fromisoformat(marriage_info['marriage_date']).year
                            if year_from <= marriage_year <= year_to:
                                marriages_found.append({
                                    "surname": surname,
                                    "husband": marriage_info.get('husband', ''),
                                    "wife": marriage_info.get('wife', ''),
                                    "date": marriage_info['marriage_date'],
                                    "year": marriage_year,
                                    "address": address,
                                    "old_address": old_address,
                                    "file_path": file_path,
                                    "type": "MAŁŻONKOWIE"
                                })
                        except Exception:
                            pass
                    
                    # Sprawdź śluby dziadków
                    gp_marriage_info = extract_grandparents_marriage_info(df)
                    gp_date = gp_marriage_info.get("marriage_date") if isinstance(gp_marriage_info, dict) else gp_marriage_info
                    if gp_date:
                        try:
                            gp_year = datetime.fromisoformat(gp_date).year
                            if year_from <= gp_year <= year_to:
                                marriages_found.append({
                                    "surname": surname,
                                    "husband": "Dziadek",
                                    "wife": "Babcia",
                                    "date": gp_date,
                                    "year": gp_year,
                                    "address": address,
                                    "old_address": old_address,
                                    "file_path": file_path,
                                    "type": "DZIADKOWIE"
                                })
                        except Exception:
                            pass
                            
                except Exception as e:
                    logging.debug(f"Błąd przetwarzania {filename}: {e}")
                    continue
            
            # Wyświetl wyniki
            results_area.delete("1.0", tk.END)
            results_area.insert(tk.END, f"Przeskanowano plików: {files_scanned}\n", "bold")
            results_area.insert(tk.END, f"Znaleziono ślubów w latach {year_from}-{year_to}: {len(marriages_found)}\n\n", "bold")
            results_area.insert(tk.END, "="*80 + "\n\n")
            
            if marriages_found:
                results_area.insert(tk.END, "Kliknij na nazwisko aby otworzyć kartotekę\n\n", "info")
                
                counter = 0
                for m in sorted(marriages_found, key=lambda x: x["year"]):
                    counter += 1
                    type_str = m.get("type", "MAŁŻONKOWIE")
                    old = m.get("old_address", "")
                    location = m.get("address", "") + (f" (stary: {old})" if old else "")
                    
                    # Wyświetl datę i rok
                    results_area.insert(tk.END, f"{m['date']} ({m['year']}) - ")
                    
                    # Link do nazwiska
                    link_tag = f"marriage_link_{counter}"
                    results_area.insert(tk.END, f"{m['surname']}: {m['husband']} i {m['wife']}", link_tag)
                    results_area.insert(tk.END, f" - {location} [{type_str}]\n")
                    
                    # Konfiguruj link
                    results_area.tag_config(link_tag, foreground="blue", underline=True)
                    if m.get("file_path") and os.path.exists(m["file_path"]):
                        results_area.tag_bind(link_tag, "<Button-1>", 
                                            lambda e, path=m["file_path"]: os.startfile(path))
                        results_area.tag_bind(link_tag, "<Enter>", 
                                            lambda e, t=link_tag: results_area.tag_config(t, font=("Courier", 10, "underline")))
                        results_area.tag_bind(link_tag, "<Leave>", 
                                            lambda e, t=link_tag: results_area.tag_config(t, font=("Courier", 10)))
            else:
                results_area.insert(tk.END, "Nie znaleziono ślubów w podanym zakresie lat.\n")
                
        except Exception as e:
            messagebox.showerror("Błąd", f"Wystąpił błąd podczas wyszukiwania:\n{e}")
    
    search_btn.config(command=search_marriages)
    year_from_entry.bind("<Return>", lambda e: search_marriages())
    year_to_entry.bind("<Return>", lambda e: search_marriages())
    
    # Przyciski na dole
    bottom_frame = tk.Frame(content, bg="#ECF0F1")
    bottom_frame.pack(fill="x", pady=(10, 0))
    
    tk.Button(bottom_frame, text="💾 Zapisz do Excel", 
             command=save_marriages_to_excel,
             bg="#27AE60", fg="white",
             font=("Segoe UI", 10, "bold"),
             relief="flat", padx=20, pady=8,
             cursor="hand2").pack(side=tk.LEFT, padx=(0, 10))
    
    tk.Button(bottom_frame, text="Zamknij", 
             command=dialog.destroy,
             bg="#95A5A6", fg="white",
             font=("Segoe UI", 10, "bold"),
             relief="flat", padx=20, pady=8,
             cursor="hand2").pack(side=tk.RIGHT)



